from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from config import DATE_FORMAT, DEFAULT_EXCLUDE_TODAY, MSK_TZ, OUTPUT_COLUMNS, OUTPUT_DIR


def build_period_dates(days: int, exclude_today: bool = DEFAULT_EXCLUDE_TODAY) -> tuple[date, date]:
    now_msk = datetime.now(MSK_TZ)
    today = now_msk.date()

    if exclude_today:
        period_start = today - timedelta(days=days)
        period_end = today - timedelta(days=1)
    else:
        period_start = today - timedelta(days=days - 1)
        period_end = today

    return period_start, period_end


def build_output_path(days: int, exclude_today: bool = DEFAULT_EXCLUDE_TODAY) -> Path:
    period_start, period_end = build_period_dates(days=days, exclude_today=exclude_today)
    file_name = (
        f"telegram_posts_{period_start.strftime('%Y-%m-%d')}_{period_end.strftime('%Y-%m-%d')}.xlsx"
    )
    return OUTPUT_DIR / file_name


def prepare_rows(items: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {
            "Название канала": str(item.get("channel_name", "")),
            "Дата публикации": str(item.get("published_at", "")),
            "Текст публикации": str(item.get("text", "")),
            "Ссылка на публикацию": str(item.get("source_url", "")),
        }
        for item in items
    ]


def save_excel(path: Path, items: list[dict[str, Any]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(prepare_rows(items), columns=OUTPUT_COLUMNS)
    df.to_excel(path, index=False)

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        'A': 28,
        'B': 22,
        'C': 100,
        'D': 45,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    workbook.save(path)

    return path


def build_metadata(items: list[dict[str, Any]], errors: list[dict[str, str]], days: int, exclude_today: bool) -> dict[str, Any]:
    period_start, period_end = build_period_dates(days=days, exclude_today=exclude_today)

    return {
        "generated_at": datetime.now(MSK_TZ).strftime(DATE_FORMAT),
        "timezone": "Europe/Moscow",
        "days": days,
        "exclude_today": exclude_today,
        "period_start": str(period_start),
        "period_end": str(period_end),
        "news_count": len(items),
        "errors_count": len(errors),
        "errors": errors,
    }
